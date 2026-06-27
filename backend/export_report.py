"""Generate a polished PDF and an Excel workbook from a meeting document.

Both rely on the same enriched meeting dict produced by routes_meetings._enrich
plus the two company display names from settings.
"""
import io

from fpdf import FPDF
from fpdf.fonts import FontFace
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BUCKETS = [("d90", "90 Days"), ("d60", "60 Days"), ("d30", "30 Days"), ("othera", "Other")]

_HEADINGS = FontFace(emphasis="BOLD", color=(0, 0, 0), fill_color=(235, 235, 235))


def _s(text):
    """Make text safe for the base PDF font (latin-1)."""
    if text is None:
        return ""
    repl = {"\u2014": "-", "\u2013": "-", "\u2019": "'", "\u201c": '"', "\u201d": '"',
            "\u20b9": "Rs ", "\u00d7": "x", "\u2192": "->", "\u2022": "-"}
    for k, v in repl.items():
        text = str(text).replace(k, v)
    return str(text).encode("latin-1", "replace").decode("latin-1")


def _amt(a):
    a = a or {}
    return (a.get("mbs", 0) or 0), (a.get("mcorp", 0) or 0)


def _inr(n):
    n = float(n or 0)
    sign = "-" if n < 0 else ""
    n = abs(n)
    if n >= 1e7:
        return f"{sign}{n/1e7:.2f} Cr"
    if n >= 1e5:
        return f"{sign}{n/1e5:.2f} L"
    return f"{sign}{n:,.0f}"


# ============================== PDF ==============================
class _PDF(FPDF):
    title_text = ""
    sub_text = ""

    def header(self):
        self.set_font("Helvetica", "B", 15)
        self.cell(0, 8, self.title_text, ln=1)
        self.set_font("Helvetica", "", 9)
        self.set_text_color(110, 110, 110)
        self.cell(0, 5, self.sub_text, ln=1)
        self.set_text_color(0, 0, 0)
        self.ln(2)

    def footer(self):
        self.set_y(-12)
        self.set_font("Helvetica", "", 7)
        self.set_text_color(150, 150, 150)
        self.cell(0, 8, f"CollectIQ  -  page {self.page_no()}", align="C")


def _section(pdf, label):
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_fill_color(20, 20, 20)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 7, f"  {label}", ln=1, fill=True)
    pdf.set_text_color(0, 0, 0)          # black body text
    pdf.set_fill_color(243, 243, 243)    # light stripe for table rows (not dark!)
    pdf.ln(1)


def build_pdf(meeting: dict, company_a: str, company_b: str) -> bytes:
    s = meeting.get("summary", {})
    pdf = _PDF(orientation="L", unit="mm", format="A4")
    pdf.title_text = _s(meeting.get("title", "Weekly Collection Meeting"))
    period = ""
    if meeting.get("period_start") or meeting.get("period_end"):
        period = f"  |  Period: {meeting.get('period_start','')} to {meeting.get('period_end','')}"
    pdf.sub_text = f"Meeting date: {meeting.get('meeting_date','')}{period}"
    pdf.set_auto_page_break(True, margin=15)
    pdf.add_page()

    # KPI strip
    pdf.set_font("Helvetica", "", 9)
    kpis = [
        ("Total Outstanding", _inr(s.get("total_outstanding"))),
        ("Collected", _inr(s.get("collected"))),
        ("Coll %", f"{s.get('coll_pct', 0):.1f}%"),
        ("Sales (Rs)", _inr(s.get("sales_value"))),
        ("Purchase (Rs)", _inr(s.get("purchase_value"))),
        ("90-day Overdue", _inr(s.get("d90"))),
    ]
    w = (pdf.w - 20) / len(kpis)
    for label, _ in kpis:
        pdf.set_font("Helvetica", "", 7)
        pdf.set_text_color(110, 110, 110)
        pdf.cell(w, 5, label, border="LTR", align="C")
    pdf.ln()
    for _, val in kpis:
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(w, 7, val, border="LBR", align="C")
    pdf.ln(8)

    # Collection table
    _section(pdf, "Collection Outstanding")
    headers = ["Rep"]
    for _, lbl in BUCKETS:
        headers += [f"{lbl} {company_a}", f"{company_b}", "Total"]
    headers += [f"Outst. {company_a}", company_b, "Total", f"Coll {company_a}", company_b, "Total", "Coll %"]
    pdf.set_font("Helvetica", "", 6)
    with pdf.table(text_align="CENTER", first_row_as_headings=True,
                   headings_style=_HEADINGS) as table:
        row = table.row()
        for h in headers:
            row.cell(h)
        for rep in meeting.get("reps", []):
            ag = rep.get("aging", {})
            o_m = o_c = 0.0
            cells = [rep.get("name", "")]
            for bkey, _ in BUCKETS:
                m, c = _amt(ag.get(bkey))
                o_m += m; o_c += c
                cells += [_inr(m), _inr(c), _inr(m + c)]
            cm, cc = _amt(rep.get("weekly_collection"))
            out = o_m + o_c
            coll = cm + cc
            pct = (coll / out * 100) if out else 0
            cells += [_inr(o_m), _inr(o_c), _inr(out), _inr(cm), _inr(cc), _inr(coll), f"{pct:.1f}%"]
            r = table.row()
            for cval in cells:
                r.cell(_s(cval))

    # Branches
    _section(pdf, "Branch Sales & Purchase (Tons)")
    pdf.set_font("Helvetica", "", 7)
    with pdf.table(text_align="CENTER", first_row_as_headings=True,
                   headings_style=_HEADINGS) as table:
        hr = table.row()
        for h in ["Branch", f"Purchase {company_a}", company_b, "Total", f"Sales {company_a}", company_b, "Total"]:
            hr.cell(h)
        for b in meeting.get("branches", []):
            pm, pc = _amt(b.get("purchase", {}).get("tons"))
            sm, sc = _amt(b.get("sales", {}).get("tons"))
            r = table.row()
            for cval in [b.get("name", ""), f"{pm:.2f}", f"{pc:.2f}", f"{pm+pc:.2f}", f"{sm:.2f}", f"{sc:.2f}", f"{sm+sc:.2f}"]:
                r.cell(_s(cval))

    # Quotation + Marketing side by side via two sections
    _section(pdf, "Quotation Pipeline")
    q = meeting.get("quotation", {})
    pdf.set_font("Helvetica", "", 7)
    with pdf.table(text_align="CENTER", first_row_as_headings=True,
                   headings_style=_HEADINGS) as table:
        hr = table.row()
        for h in ["Stage", company_a, company_b, "Total"]:
            hr.cell(h)
        for key, lbl in [("prepair", "Prepared"), ("conform", "Confirmed"), ("pending", "Pending"),
                         ("under_process", "Under Process"), ("not_conform", "Not Confirmed")]:
            m, c = _amt(q.get(key))
            r = table.row()
            for cval in [lbl, f"{m:.0f}", f"{c:.0f}", f"{m+c:.0f}"]:
                r.cell(_s(cval))

    _section(pdf, "Marketing Activity")
    pdf.set_font("Helvetica", "", 7)
    with pdf.table(text_align="CENTER", first_row_as_headings=True,
                   headings_style=_HEADINGS) as table:
        hr = table.row()
        for h in ["Person", "Visits", "Inquiries", "Inq. Confirmed", "Order Loss", "Target Tons", "Achieve %"]:
            hr.cell(h)
        for mrep in meeting.get("marketing_reps", []):
            vm, vc = _amt(mrep.get("visit"))
            im, ic = _amt(mrep.get("inquiry"))
            cfm, cfc = _amt(mrep.get("inquiry_conform"))
            lm, lc = _amt(mrep.get("order_loss"))
            r = table.row()
            for cval in [mrep.get("name", ""), f"{vm+vc:.0f}", f"{im+ic:.0f}", f"{cfm+cfc:.0f}",
                         f"{lm+lc:.0f}", f"{mrep.get('target_tons', 0):.0f}", f"{mrep.get('target_tons_achieve_pct', 0):.0f}%"]:
                r.cell(_s(cval))

    out = pdf.output()
    return bytes(out)


# ============================== Excel ==============================
_HEAD = Font(bold=True, color="FFFFFF")
_HEAD_FILL = PatternFill("solid", fgColor="141414")
_BOLD = Font(bold=True)
_CENTER = Alignment(horizontal="center")
_thin = Side(style="thin", color="DDDDDD")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = _HEAD
        c.fill = _HEAD_FILL
        c.alignment = _CENTER
        c.border = _BORDER


def build_xlsx(meeting: dict, company_a: str, company_b: str) -> bytes:
    s = meeting.get("summary", {})
    wb = Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = meeting.get("title", "Weekly Collection Meeting")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Meeting date: {meeting.get('meeting_date','')}    Period: {meeting.get('period_start','')} to {meeting.get('period_end','')}"
    rows = [
        ("Total Outstanding", s.get("total_outstanding")),
        ("90 Days", s.get("d90")), ("60 Days", s.get("d60")),
        ("30 Days", s.get("d30")), ("Other", s.get("othera")),
        ("Collected this week", s.get("collected")),
        ("Collection %", s.get("coll_pct")),
        ("New Target", s.get("new_target_total")),
        ("Sales (Rs)", s.get("sales_value")), ("Purchase (Rs)", s.get("purchase_value")),
        ("Sales (Tons)", s.get("sales_tons")), ("Purchase (Tons)", s.get("purchase_tons")),
    ]
    r0 = 4
    ws.cell(row=r0, column=1, value="Metric").font = _BOLD
    ws.cell(row=r0, column=2, value="Value").font = _BOLD
    for i, (k, v) in enumerate(rows, start=r0 + 1):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=round(float(v or 0), 2))
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18

    # Collection sheet
    wc = wb.create_sheet("Collection")
    head = ["Rep"]
    for _, lbl in BUCKETS:
        head += [f"{lbl} {company_a}", f"{lbl} {company_b}", f"{lbl} Total"]
    head += [f"Outstanding {company_a}", f"Outstanding {company_b}", "Outstanding Total",
             f"Collected {company_a}", f"Collected {company_b}", "Collected Total", "Coll %"]
    wc.append(head)
    _style_header(wc, 1, len(head))
    for rep in meeting.get("reps", []):
        ag = rep.get("aging", {})
        o_m = o_c = 0.0
        row = [rep.get("name", "")]
        for bkey, _ in BUCKETS:
            m, c = _amt(ag.get(bkey))
            o_m += m; o_c += c
            row += [m, c, m + c]
        cm, cc = _amt(rep.get("weekly_collection"))
        out = o_m + o_c
        coll = cm + cc
        row += [o_m, o_c, out, cm, cc, coll, round((coll / out * 100) if out else 0, 1)]
        wc.append(row)
    for col in wc.columns:
        wc.column_dimensions[col[0].column_letter].width = 15

    # Branches
    wbr = wb.create_sheet("Branches")
    bh = ["Branch", f"Purchase {company_a} (T)", f"Purchase {company_b} (T)", "Purchase Total (T)",
          f"Sales {company_a} (T)", f"Sales {company_b} (T)", "Sales Total (T)"]
    wbr.append(bh)
    _style_header(wbr, 1, len(bh))
    for b in meeting.get("branches", []):
        pm, pc = _amt(b.get("purchase", {}).get("tons"))
        sm, sc = _amt(b.get("sales", {}).get("tons"))
        wbr.append([b.get("name", ""), pm, pc, pm + pc, sm, sc, sm + sc])
    for col in wbr.columns:
        wbr.column_dimensions[col[0].column_letter].width = 18

    # Quotation
    wq = wb.create_sheet("Quotation")
    wq.append(["Stage", company_a, company_b, "Total"])
    _style_header(wq, 1, 4)
    q = meeting.get("quotation", {})
    for key, lbl in [("prepair", "Prepared"), ("conform", "Confirmed"), ("pending", "Pending"),
                     ("under_process", "Under Process"), ("not_conform", "Not Confirmed")]:
        m, c = _amt(q.get(key))
        wq.append([lbl, m, c, m + c])
    for col in wq.columns:
        wq.column_dimensions[col[0].column_letter].width = 16

    # Marketing
    wm = wb.create_sheet("Marketing")
    mh = ["Person", f"Visits {company_a}", f"Visits {company_b}", "Inquiries", "Inq. Confirmed",
          "Order Loss", "Target Tons", "Achieve %", "Target Party", "Achieve %"]
    wm.append(mh)
    _style_header(wm, 1, len(mh))
    for mrep in meeting.get("marketing_reps", []):
        vm, vc = _amt(mrep.get("visit"))
        im, ic = _amt(mrep.get("inquiry"))
        cfm, cfc = _amt(mrep.get("inquiry_conform"))
        lm, lc = _amt(mrep.get("order_loss"))
        wm.append([mrep.get("name", ""), vm, vc, im + ic, cfm + cfc, lm + lc,
                   mrep.get("target_tons", 0), mrep.get("target_tons_achieve_pct", 0),
                   mrep.get("target_party", 0), mrep.get("target_party_achieve_pct", 0)])
    for col in wm.columns:
        wm.column_dimensions[col[0].column_letter].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================== Trends report PDF (all departments) ==============================
def build_trends_pdf(weeks: list, company_a: str, company_b: str) -> bytes:
    pdf = _PDF(orientation="L", unit="mm", format="A4")
    pdf.title_text = "Trends & Analytics Report"
    if weeks:
        pdf.sub_text = f"All departments  |  {weeks[0]['date']} to {weeks[-1]['date']}  |  {len(weeks)} week(s)"
    pdf.set_auto_page_break(True, margin=15)
    pdf.add_page()

    if not weeks:
        pdf.set_font("Helvetica", "", 11)
        pdf.cell(0, 10, "No meeting data yet.", ln=1)
        return bytes(pdf.output())

    # ---- Department 1: Collection ----
    _section(pdf, "1. Collection")
    pdf.set_font("Helvetica", "", 7)
    with pdf.table(text_align="CENTER", first_row_as_headings=True, headings_style=_HEADINGS) as t:
        hr = t.row()
        for h in ["Week", "Outstanding", "90-day", "60-day", "30-day", "Other", "Collected", "Coll %"]:
            hr.cell(h)
        for w in weeks:
            r = t.row()
            for v in [w["date"], _inr(w["outstanding"]), _inr(w["d90"]), _inr(w["d60"]), _inr(w["d30"]),
                      _inr(w["othera"]), _inr(w["collected"]), f"{w['coll_pct']:.1f}%"]:
                r.cell(_s(v))

    # ---- Department 2: Sales vs Purchase ----
    _section(pdf, "2. Sales vs Purchase")
    pdf.set_font("Helvetica", "", 7)
    with pdf.table(text_align="CENTER", first_row_as_headings=True, headings_style=_HEADINGS) as t:
        hr = t.row()
        for h in ["Week", "Sales (Rs)", "Purchase (Rs)", "Sales (T)", "Purchase (T)", "Net (Rs)"]:
            hr.cell(h)
        for w in weeks:
            r = t.row()
            net = w["sales_value"] - w["purchase_value"]
            for v in [w["date"], _inr(w["sales_value"]), _inr(w["purchase_value"]),
                      f"{w['sales_tons']:.2f}", f"{w['purchase_tons']:.2f}", _inr(net)]:
                r.cell(_s(v))

    # Branchwise (latest week)
    latest = weeks[-1]
    if latest.get("branches"):
        pdf.ln(1)
        pdf.set_font("Helvetica", "I", 7)
        pdf.cell(0, 4, f"Branchwise - latest week ({latest['date']}):", ln=1)
        pdf.set_font("Helvetica", "", 7)
        with pdf.table(text_align="CENTER", first_row_as_headings=True, headings_style=_HEADINGS) as t:
            hr = t.row()
            for h in ["Branch", "Sales (T)", "Purchase (T)"]:
                hr.cell(h)
            for b in latest["branches"]:
                r = t.row()
                for v in [b["name"], f"{b['sales_tons']:.2f}", f"{b['purchase_tons']:.2f}"]:
                    r.cell(_s(v))

    # ---- Department 3: Marketing ----
    _section(pdf, "3. Marketing")
    pdf.set_font("Helvetica", "", 7)
    with pdf.table(text_align="CENTER", first_row_as_headings=True, headings_style=_HEADINGS) as t:
        hr = t.row()
        for h in ["Week", "Visits", "Inquiries", "Confirmed", "Order Loss", "Conversion %"]:
            hr.cell(h)
        for w in weeks:
            r = t.row()
            for v in [w["date"], f"{w['visits']:.0f}", f"{w['inquiries']:.0f}", f"{w['confirmed']:.0f}",
                      f"{w['order_loss']:.0f}", f"{w['conv']:.1f}%"]:
                r.cell(_s(v))

    # ---- Analysis ----
    _section(pdf, "4. Analysis")
    pdf.set_font("Helvetica", "", 9)
    n = len(weeks)
    avg_pct = sum(w["coll_pct"] for w in weeks) / n
    tot_coll = sum(w["collected"] for w in weeks)
    tot_sales = sum(w["sales_value"] for w in weeks)
    tot_purch = sum(w["purchase_value"] for w in weeks)
    first, last = weeks[0], weeks[-1]
    out_delta = ((last["outstanding"] - first["outstanding"]) / first["outstanding"] * 100) if first["outstanding"] else 0
    best = max(weeks, key=lambda w: w["coll_pct"])
    lines = [
        f"- Average collection rate over the period: {avg_pct:.1f}%.",
        f"- Total collected: {_inr(tot_coll)}  |  Total sales: {_inr(tot_sales)}  |  Total purchase: {_inr(tot_purch)}.",
        f"- Outstanding moved {out_delta:+.1f}% from {first['date']} to {last['date']}.",
        f"- Best collection week: {best['date']} at {best['coll_pct']:.1f}%.",
        f"- Net sales-purchase position (latest): {_inr(last['sales_value'] - last['purchase_value'])}.",
    ]
    for ln_ in lines:
        pdf.cell(0, 6, ln_, ln=1)

    return bytes(pdf.output())
